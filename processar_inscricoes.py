import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STATUS_ATIVOS = {'pago', 'deferida', 'deferido'}

def _borda(cor='000000', estilo='thin'):
    s = Side(border_style=estilo, color=cor)
    return Border(top=s, bottom=s, left=s, right=s)

def _borda_leve():
    s = Side(border_style='thin', color='CCCCCC')
    return Border(top=s, bottom=s, left=s, right=s)

def _header_style():
    return {
        'font': Font(bold=True, color='FFFFFF', name='Arial', size=9),
        'fill': PatternFill('solid', fgColor='1F4E8C'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _borda()
    }

def _aplicar(cell, **kwargs):
    for k,v in kwargs.items():
        setattr(cell, k, v)

CORES_CARGO = ['FFFFFF', 'EBF3FF']

def _criar_aba(wb, nome_aba, titulo, colunas, larguras, dados, nome_concurso):
    ws = wb.create_sheet(nome_aba)
    NC = len(colunas)

    # Linha 1: título mesclado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c = ws.cell(row=1, column=1, value=f"{nome_concurso}\n\nRESULTADO PRELIMINAR DAS INSCRIÇÕES - {titulo}")
    _aplicar(c,
        font=Font(bold=True, color='FFFFFF', name='Arial', size=10),
        fill=PatternFill('solid', fgColor='1F4E8C'),
        alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
        border=_borda()
    )
    ws.row_dimensions[1].height = 50

    # Linha 2: cabeçalho
    hs = _header_style()
    for ci, col in enumerate(colunas, 1):
        c = ws.cell(row=2, column=ci, value=col)
        _aplicar(c, **hs)
        ws.column_dimensions[get_column_letter(ci)].width = larguras.get(col, 14)
    ws.row_dimensions[2].height = 28

    # Dados
    cargo_atual = ''
    cor_idx = 0
    for ri, row in enumerate(dados, 3):
        cargo = str(row.get('CARGO', ''))
        if cargo != cargo_atual:
            cargo_atual = cargo
            cor_idx = (cor_idx + 1) % 2
        fill = PatternFill('solid', fgColor=CORES_CARGO[cor_idx])

        for ci, col in enumerate(colunas, 1):
            val = row.get(col, '')
            if pd.isna(val) or str(val) in ('nan', 'NaT', '<NA>'):
                val = ''
            c = ws.cell(row=ri, column=ci, value=str(val) if val != '' else '')
            _aplicar(c,
                font=Font(name='Arial', size=9),
                fill=fill,
                alignment=Alignment(vertical='center'),
                border=_borda_leve()
            )
        ws.row_dimensions[ri].height = 14

    ws.freeze_panes = 'A3'
    return ws

LARGURAS = {
    'INSCRIÇÃO': 12, 'CANDIDATO': 35, 'CPF': 14,
    'DATA DE NASCIMENTO': 12, 'CARGO': 45,
    'STATUS DE PAGAMENTO': 16, 'CONCORRE VAGA DEFICIÊNCIA': 10,
    'DEFICIÊNCIA': 22, 'CONCORRE VAGA NEGRO': 10,
    'PRECISA ATENDIMENTO ESPECIAL': 10, 'ATENDIMENTO ESPECIAL': 25,
    'DATA REALIZAÇÃO PROVA': 18, 'RESULTADO': 12,
    'JUSTIFICATIVA': 40, 'LINK': 35,
}

def processar(arquivo_bytes, conjuntos_mesma_prova=None, processo=None):
    # Ler arquivo
    df = pd.read_excel(io.BytesIO(arquivo_bytes), header=1, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].str.strip()

    # Filtrar processo
    col_proc = next((c for c in df.columns if any(x in c.upper() for x in ['CONCURSO','EDITAL','PROCESSO'])), None)
    if col_proc and processo:
        df = df[df[col_proc].str.strip() == processo.strip()]

    nome_concurso = processo or (df[col_proc].iloc[0] if col_proc and len(df) else 'CONCURSO')

    # Normalizar campos
    df['_cpf'] = df['CPF'].str.replace(r'\D','',regex=True)
    df['_cod'] = df.get('CÓDIGO', df.get('CODIGO', pd.Series([''] * len(df)))).fillna('')
    df['_status'] = df['STATUS'].str.lower().str.strip()
    df['_dt'] = pd.to_datetime(df.get('DATA INSCRIÇÃO', df.get('DATA INSCRICAO', '')), errors='coerce')

    cancelar = set()

    # Regra 1: duplicata no mesmo cargo
    for (cpf, cod), grp in df.groupby(['_cpf','_cod'], dropna=False):
        ativos = grp[grp['_status'].isin(STATUS_ATIVOS)].sort_values('_dt', ascending=True)
        cancelar.update(ativos.iloc[:-1]['INSCRIÇÃO'].tolist())

    # Regra 2: conjuntos de mesma prova
    if conjuntos_mesma_prova:
        for conj in conjuntos_mesma_prova:
            sub = df[df['_cod'].isin([str(c).strip() for c in conj])]
            for cpf, grp in sub.groupby('_cpf', dropna=False):
                ativos = grp[grp['_status'].isin(STATUS_ATIVOS)].sort_values('_dt', ascending=True)
                cancelar.update(ativos.iloc[:-1]['INSCRIÇÃO'].tolist())

    df.loc[df['INSCRIÇÃO'].isin(cancelar), 'STATUS'] = 'Cancelada'
    df['_status'] = df['STATUS'].str.lower().str.strip()

    # Normalizar datas para exibição
    def fmt_data(v):
        try:
            d = pd.to_datetime(v, errors='coerce')
            return d.strftime('%d/%m/%Y') if not pd.isna(d) else str(v) if v else ''
        except:
            return str(v) if v else ''

    col_nasc = next((c for c in df.columns if 'NASC' in c.upper()), None)
    col_prova = next((c for c in df.columns if 'REALIZAÇÃO' in c.upper() or 'DATA PROVA' in c.upper()), None)
    if col_nasc: df['DATA DE NASCIMENTO'] = df[col_nasc].apply(fmt_data)
    if col_prova: df['DATA REALIZAÇÃO PROVA'] = df[col_prova].apply(fmt_data)

    df['STATUS DE PAGAMENTO'] = df['STATUS']
    df = df.sort_values(['CARGO', 'INSCRIÇÃO'])

    # ── Gerar planilha de resultado ────────────────────────────────────
    wb_res = Workbook()
    wb_res.remove(wb_res.active)

    COLS_AMPLA = ['INSCRIÇÃO','CANDIDATO','CPF','DATA DE NASCIMENTO','CARGO',
        'STATUS DE PAGAMENTO','CONCORRE VAGA DEFICIÊNCIA','DEFICIÊNCIA',
        'CONCORRE VAGA NEGRO','PRECISA ATENDIMENTO ESPECIAL','ATENDIMENTO ESPECIAL',
        'DATA REALIZAÇÃO PROVA']
    cols_ampla_ok = [c for c in COLS_AMPLA if c in df.columns or c == 'STATUS DE PAGAMENTO' or c == 'DATA DE NASCIMENTO']

    def to_rows(sub, cols):
        rows = []
        for _, r in sub.iterrows():
            row = {}
            for c in cols:
                row[c] = r.get(c, '')
            rows.append(row)
        return rows

    _criar_aba(wb_res, 'AMPLA', 'AMPLA', cols_ampla_ok, LARGURAS, to_rows(df, cols_ampla_ok), nome_concurso)

    # PCD
    pcd = df[(df.get('CONCORRE VAGA DEFICIÊNCIA','').str.upper()=='SIM') & df['_status'].isin(STATUS_ATIVOS)] if 'CONCORRE VAGA DEFICIÊNCIA' in df.columns else pd.DataFrame()
    if len(pcd):
        cols_pcd = [c for c in ['INSCRIÇÃO','CANDIDATO','CPF','DATA DE NASCIMENTO','CARGO',
            'STATUS DE PAGAMENTO','CONCORRE VAGA DEFICIÊNCIA','DEFICIÊNCIA',
            'RESULTADO PCD','JUSTIFICATIVA PCD','LINK PCD'] if c in df.columns or c in ['DATA DE NASCIMENTO','STATUS DE PAGAMENTO']]
        _criar_aba(wb_res, 'PCD', 'PcD', cols_pcd, LARGURAS, to_rows(pcd, cols_pcd), nome_concurso)

    # PNP
    pnp = df[(df.get('CONCORRE VAGA NEGRO','').str.upper()=='SIM') & df['_status'].isin(STATUS_ATIVOS)] if 'CONCORRE VAGA NEGRO' in df.columns else pd.DataFrame()
    if len(pnp):
        cols_pnp = [c for c in ['INSCRIÇÃO','CANDIDATO','CPF','DATA DE NASCIMENTO','CARGO',
            'STATUS DE PAGAMENTO','CONCORRE VAGA NEGRO','DATA REALIZAÇÃO PROVA'] if c in df.columns or c in ['DATA DE NASCIMENTO','STATUS DE PAGAMENTO']]
        _criar_aba(wb_res, 'PNP', 'PNP', cols_pnp, LARGURAS, to_rows(pnp, cols_pnp), nome_concurso)

    # COND. ESPECIAL
    cond_col = next((c for c in df.columns if 'ATENDIMENTO ESPECIAL' in c.upper() and 'PRECISA' in c.upper()), None)
    cond = df[(df[cond_col].str.upper()=='SIM') & df['_status'].isin(STATUS_ATIVOS)] if cond_col else pd.DataFrame()
    if len(cond):
        cols_cond = [c for c in ['INSCRIÇÃO','CANDIDATO','CPF','DATA DE NASCIMENTO','CARGO',
            'STATUS DE PAGAMENTO','ATENDIMENTO ESPECIAL',
            'RESULTADO COND. ESPECIAL','JUSTIFICATIVA COND. ESPECIAL','LINK COND. ESPECIAL',
            'DATA REALIZAÇÃO PROVA'] if c in df.columns or c in ['DATA DE NASCIMENTO','STATUS DE PAGAMENTO']]
        _criar_aba(wb_res, 'COND.ESPECIAL', 'CONDIÇÃO ESPECIAL', cols_cond, LARGURAS, to_rows(cond, cols_cond), nome_concurso)

    buf_res = io.BytesIO()
    wb_res.save(buf_res)

    # ── Gerar planilha de alocação ─────────────────────────────────────
    wb_aloc = Workbook()
    wb_aloc.remove(wb_aloc.active)

    # Aba ALOCAÇÃO: pagos e deferidos sem cancelados
    aloc = df[df['_status'].isin(STATUS_ATIVOS)]
    _criar_aba(wb_aloc, 'ALOCAÇÃO', 'ALOCAÇÃO', cols_ampla_ok, LARGURAS, to_rows(aloc, cols_ampla_ok), nome_concurso)

    # Aba CANCELADOS POR CONJUNTO: candidatos removidos pela regra 2
    if conjuntos_mesma_prova and cancelar:
        cancel_df = df[df['INSCRIÇÃO'].isin(cancelar)]
        # Só os que foram cancelados pela regra de conjuntos
        cancel_por_conj = []
        for conj in conjuntos_mesma_prova:
            sub = cancel_df[cancel_df['_cod'].isin([str(c).strip() for c in conj])]
            for _, r in sub.iterrows():
                cancel_por_conj.append({**dict(r), 'CONJUNTO': ', '.join(conj)})
        if cancel_por_conj:
            cols_cancel = ['INSCRIÇÃO','CANDIDATO','CPF','CARGO','STATUS DE PAGAMENTO','CONJUNTO']
            _criar_aba(wb_aloc, 'CANCELADOS CONJUNTO', 'CANCELADOS - MESMA PROVA',
                cols_cancel, LARGURAS, cancel_por_conj, nome_concurso)

    buf_aloc = io.BytesIO()
    wb_aloc.save(buf_aloc)

    resumo = {
        'total': len(df),
        'pagos': (df['_status']=='pago').sum(),
        'deferidos': df['_status'].isin({'deferida','deferido'}).sum(),
        'pendentes': (df['_status']=='pendente').sum(),
        'indeferidos': (df['_status']=='indeferida').sum(),
        'cancelados': (df['_status']=='cancelada').sum(),
        'total_alocacao': len(aloc),
        'abas_resultado': wb_res.sheetnames,
    }

    return buf_res.getvalue(), buf_aloc.getvalue(), resumo

def df_para_xlsx(df, titulo):
    """Mantido para compatibilidade"""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=col)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
