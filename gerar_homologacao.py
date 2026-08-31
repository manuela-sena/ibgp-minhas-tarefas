import pandas as pd
import io
import re
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Mapeamento de reservas: nome da aba → coluna de classificação na CF → coluna flag na CF
RESERVAS = [
    ("AMPLA",              "CLASS. CARGO AMPLA",        "AMPLA"),
    ("PCD",                "CLASS. CARGO PcD",          "PCD.1"),
    ("PNP",                "CLASS. CARGO PNP",          "PN4"),
    ("INDIGENA",           "CLASS. INDIGENA",           "IND5"),
    ("QUILOMBOLA",         "CLASS. QUILOMBOLA",         "QUILO6"),
    ("VAGAS AFIRMATIVAS",  "CLASS. VAGAS AFIRMATIVAS",  "VAGAFIRM7"),
]

def ordem_class(val):
    """Converte '1º', '2º', etc para int para ordenação."""
    try:
        return int(str(val).replace('º','').replace('ª','').strip())
    except:
        return 9999

def borda():
    s = Side(border_style='thin', color='000000')
    return Border(top=s, bottom=s, left=s, right=s)

def _norm_col(s):
    """Normaliza nome de coluna pra comparação: maiúsculo, sem acento, sem
    quebra de linha/espaço duplicado, e sem a palavra 'CARGO' (alguns editais
    usam 'CLASS. CARGO PcD', outros só 'CLASS. PcD' — é a mesma coisa)."""
    s = re.sub(r'\s+', ' ', str(s)).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.replace('CARGO ', '')
    return re.sub(r'\s+', ' ', s).strip()

def _ler_classificacao_final(bytes_cf):
    """Lê a Classificação Final. Alguns editais exportam tudo numa aba única
    chamada 'CONVOCADOS'; outros exportam uma aba por grupo de escolaridade/
    cargo (ex: FUNDAMENTAL, MÉDIO, SUPERIOR, CARGO 205, 537...) — cada uma
    com as mesmas colunas CLASS. AMPLA / CLASS. PcD / etc. Detecta os dois
    formatos automaticamente e junta tudo num único DataFrame."""
    todas_abas = pd.read_excel(io.BytesIO(bytes_cf), sheet_name=None, header=1, dtype=str)
    if 'CONVOCADOS' in todas_abas:
        abas = {'CONVOCADOS': todas_abas['CONVOCADOS']}
    else:
        abas = todas_abas

    partes = []
    for df in abas.values():
        if df is None or df.empty:
            continue
        df = df.copy()
        cols = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]
        # Algumas abas repetem o mesmo título de coluna (ex: 'INÍCIO' aparece
        # duas vezes, uma pra cada prazo) — isso quebra o pd.concat mais na
        # frente, então desambigua mantendo a primeira ocorrência com o nome
        # original e sufixando as repetidas.
        vistos = {}
        cols_unicas = []
        for c in cols:
            if c not in vistos:
                vistos[c] = 0
                cols_unicas.append(c)
            else:
                vistos[c] += 1
                cols_unicas.append(f"{c}__{vistos[c]}")
        df.columns = cols_unicas
        if 'INSCRIÇÃO' not in df.columns:
            continue  # aba sem dados de candidato (ex: aba de instruções)
        partes.append(df)

    if not partes:
        raise ValueError("Não encontrei nenhuma aba com coluna INSCRIÇÃO na Classificação Final.")

    df_cf = pd.concat(partes, ignore_index=True, sort=False)
    df_cf['INSCRIÇÃO'] = df_cf['INSCRIÇÃO'].astype(str).str.strip()
    return df_cf

def processar_homologacao(bytes_dados, bytes_cf):
    """
    bytes_dados: planilha de dados gerais (qualquer aba usada como base)
    bytes_cf: planilha de classificação final (aba CONVOCADOS, header na linha 2)
    Retorna: bytes do xlsx de homologação
    """
    # ── Ler classificação final (aba única 'CONVOCADOS' ou uma aba por
    #    escolaridade/cargo — ver _ler_classificacao_final) ────────────
    df_cf = _ler_classificacao_final(bytes_cf)
    # mapa nome-normalizado -> nome real da coluna, pra achar a coluna certa
    # mesmo com variações de acento/maiúscula/'CARGO'/quebra de linha entre editais
    cf_cols_norm = {}
    for c in df_cf.columns:
        cf_cols_norm.setdefault(_norm_col(c), c)

    # ── Ler planilha de dados (primeira aba disponível como template) ─
    xl_dados = pd.read_excel(io.BytesIO(bytes_dados), sheet_name=None, header=0, dtype=str)
    # Usar primeira aba para pegar as colunas corretas
    primeira_aba = list(xl_dados.values())[0]
    colunas = list(primeira_aba.columns)

    # Concatenar todas as abas de dados (para ter todos os candidatos)
    df_todos = pd.concat(xl_dados.values(), ignore_index=True)
    df_todos['INSCRIÇÃO'] = df_todos['INSCRIÇÃO'].astype(str).str.strip()

    # ── Criar workbook de saída ──────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remover aba padrão

    for nome_aba, col_class, col_flag in RESERVAS:
        # Filtrar candidatos desta reserva na CF — busca a coluna pelo nome
        # normalizado, porque editais diferentes escrevem o cabeçalho de
        # jeitos ligeiramente diferentes (ex: 'CLASS. CARGO PcD' vs
        # 'CLASS. PcD', com ou sem quebra de linha, com ou sem acento).
        col_real = cf_cols_norm.get(_norm_col(col_class))
        if col_real is None:
            continue

        # Candidatos que têm classificação nesta reserva (não nulo e não '-')
        mask = df_cf[col_real].notna() & (~df_cf[col_real].isin(['-','NaN','nan','']))
        df_reserva_cf = df_cf[mask][['INSCRIÇÃO', col_real, 'CARGO']].copy()
        
        if df_reserva_cf.empty:
            continue

        df_reserva_cf.columns = ['INSCRIÇÃO', 'CLASS.', 'CARGO_CF']
        df_reserva_cf['_ordem'] = df_reserva_cf['CLASS.'].apply(ordem_class)

        # PROCV: juntar com dados gerais pelo número de inscrição
        df_merged = df_reserva_cf.merge(
            df_todos, on='INSCRIÇÃO', how='left', suffixes=('_cf','')
        )

        # Atualizar classificação com a da CF
        df_merged['CLASS.'] = df_merged['CLASS._cf'] if 'CLASS._cf' in df_merged.columns else df_merged['CLASS.']

        # Ordenar por CÓDIGO do cargo e depois por classificação
        if 'CÓDIGO' in df_merged.columns:
            df_merged['_cod_sort'] = pd.to_numeric(df_merged['CÓDIGO'], errors='coerce').fillna(9999)
            df_merged = df_merged.sort_values(['_cod_sort','_ordem']).reset_index(drop=True)
        else:
            df_merged = df_merged.sort_values(['_ordem']).reset_index(drop=True)

        # Renumerar classificação sequencial por cargo
        if 'CARGO' in df_merged.columns:
            df_merged['CLASS.'] = df_merged.groupby('CÓDIGO' if 'CÓDIGO' in df_merged.columns else 'CARGO_CF').cumcount() + 1
        else:
            df_merged['CLASS.'] = range(1, len(df_merged)+1)

        # Selecionar só colunas da planilha de dados original
        cols_saida = [c for c in colunas if c in df_merged.columns]
        df_saida = df_merged[cols_saida].copy()

        # ── Criar aba ────────────────────────────────────────────────
        ws = wb.create_sheet(title=nome_aba)

        # Cabeçalho
        for ci, col in enumerate(cols_saida, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1F4E8C')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = borda()
        ws.row_dimensions[1].height = 25

        # Larguras de coluna
        col_widths = {
            'CLASS.': 8, 'INSCRIÇÃO': 12, 'ID CANDIDATO': 12,
            'CANDIDATO': 35, 'CPF': 15, 'RG': 12,
            'CÓDIGO': 8, 'CARGO': 40, 'STATUS': 10,
            'DATA NASCIMENTO': 14, 'SEXO': 6, 'EMAIL': 30,
            'CELULAR': 14, 'TELEFONE': 14,
        }
        for ci, col in enumerate(cols_saida, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 14)

        # Dados
        CORES_CARGO = ['DEEAF1', 'E2EFDA']  # alternância por cargo
        cargo_atual = None
        cor_idx = 0

        for ri, row in enumerate(df_saida.itertuples(index=False), 2):
            # Detectar mudança de cargo para alternar cor
            cargo_val = getattr(row, 'CARGO', None) or getattr(row, 'CARGO_CF', None)
            if cargo_val != cargo_atual:
                cargo_atual = cargo_val
                cor_idx = (cor_idx + 1) % 2

            fill = PatternFill('solid', fgColor=CORES_CARGO[cor_idx])

            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci)
                # Tratar NaN
                if pd.isna(val) or str(val) in ('nan','NaT','<NA>'):
                    c.value = ''
                else:
                    c.value = str(val) if not isinstance(val, (int, float)) else val
                c.font = Font(name='Arial', size=9)
                c.alignment = Alignment(vertical='center')
                c.border = borda()
                c.fill = fill
            ws.row_dimensions[ri].height = 14

        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_saida))}1"

    if not wb.sheetnames:
        ws = wb.create_sheet('SEM_DADOS')
        ws['A1'] = 'Nenhum dado encontrado.'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
