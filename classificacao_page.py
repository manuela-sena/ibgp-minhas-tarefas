# -*- coding: utf-8 -*-
"""Página nativa Streamlit: monta a classificação final de um concurso a
partir de várias planilhas (inscrições, objetiva, discursiva, títulos,
heteroidentificação, av. médica/psicológica etc.), usando o motor genérico
em motor_classificacao.py. Cada concurso configura suas próprias etapas,
cortes e critérios de desempate — a configuração pode ser salva/carregada
como JSON no próprio repositório do GitHub.
"""
import io
import json
import base64
import datetime

import streamlit as st
import openpyxl

from motor_classificacao import MotorClassificacao, is_num

REPO = "manuela-sena/ibgp-minhas-tarefas"
CONFIG_DIR = "configs_classificacao"

TIPOS_ETAPA = ["pontuacao", "classificatoria", "binaria"]
TIPOS_DESEMPATE = ["idoso_60", "maior_nota", "jurado", "idade_maior"]
POOLS_DISPONIVEIS = ["pcd", "pnp", "indigena", "quilombola"]


# ── util: github (config) ──────────────────────────────────────────────

def _gh_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def listar_configs(gh_token):
    import requests
    r = requests.get(f"https://api.github.com/repos/{REPO}/contents/{CONFIG_DIR}",
                      headers=_gh_headers(gh_token))
    if r.status_code != 200:
        return []
    return [it["name"][:-5] for it in r.json() if it["name"].endswith(".json")]


def carregar_config(gh_token, nome):
    import requests
    r = requests.get(f"https://api.github.com/repos/{REPO}/contents/{CONFIG_DIR}/{nome}.json",
                      headers=_gh_headers(gh_token))
    if r.status_code != 200:
        return None
    conteudo = base64.b64decode(r.json()["content"]).decode("utf-8")
    return json.loads(conteudo)


def salvar_config(gh_token, nome, config):
    import requests
    path = f"{CONFIG_DIR}/{nome}.json"
    r_sha = requests.get(f"https://api.github.com/repos/{REPO}/contents/{path}",
                         headers=_gh_headers(gh_token))
    sha = r_sha.json().get("sha") if r_sha.status_code == 200 else None
    conteudo = json.dumps(config, ensure_ascii=False, indent=2)
    payload = {"message": f"chore: config classificação {nome}",
               "content": base64.b64encode(conteudo.encode()).decode()}
    if sha:
        payload["sha"] = sha
    return requests.put(f"https://api.github.com/repos/{REPO}/contents/{path}",
                        headers=_gh_headers(gh_token), json=payload)


# ── util: leitura de planilhas ─────────────────────────────────────────

def ler_planilha(arquivo_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), read_only=True, data_only=True)
    return wb


def detectar_colunas(wb, sheet, linha_cabecalho):
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=linha_cabecalho, values_only=True), 1):
        if i == linha_cabecalho:
            return [c for c in row if c is not None]
    return []


def ler_linhas(wb, sheet, linha_cabecalho, colunas_map):
    """colunas_map: {"campo_logico": "Nome exato da coluna no cabeçalho" ou None}
    Retorna lista de dicts {campo_logico: valor} por linha de dados."""
    ws = wb[sheet]
    header = None
    linhas = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i < linha_cabecalho:
            continue
        if i == linha_cabecalho:
            header = list(row)
            idx = {}
            for campo, nome_col in colunas_map.items():
                if nome_col and nome_col in header:
                    idx[campo] = header.index(nome_col)
            continue
        if row is None or row[0] is None:
            continue
        item = {}
        for campo, pos in idx.items():
            item[campo] = row[pos] if pos < len(row) else None
        linhas.append(item)
    return linhas


# ── UI principal ────────────────────────────────────────────────────────

def _config_vazia():
    return {
        "nome": "",
        "pools": list(POOLS_DISPONIVEIS),
        "desempate": [],
        "etapas": [],
    }


def render(gh_token):
    st.markdown("""
<style>
header[data-testid="stHeader"]{display:none}
[data-testid="stAppViewContainer"]{background:#eef1f6}
.block-container{padding:1.5rem 2rem!important}
</style>""", unsafe_allow_html=True)

    col_back, col_title = st.columns([1, 8])
    with col_back:
        if st.button("← Voltar"):
            st.session_state.pop("pagina", None)
            st.rerun()
    with col_title:
        st.markdown("## 🏆 Classificação Final")

    st.markdown("---")

    if "clf_config" not in st.session_state:
        st.session_state["clf_config"] = _config_vazia()
    config = st.session_state["clf_config"]

    # ---- carregar/salvar config ----
    with st.expander("📁 Configuração do concurso", expanded=True):
        configs_salvas = listar_configs(gh_token) if gh_token else []
        c1, c2 = st.columns([3, 1])
        with c1:
            escolha = st.selectbox("Carregar configuração salva", [""] + configs_salvas,
                                    format_func=lambda x: "Nova configuração..." if x == "" else x)
        with c2:
            st.write("")
            if st.button("Carregar", use_container_width=True) and escolha:
                carregada = carregar_config(gh_token, escolha)
                if carregada:
                    st.session_state["clf_config"] = carregada
                    st.rerun()

        config["nome"] = st.text_input("Nome do concurso", value=config.get("nome", ""),
                                        placeholder="Ex: Betim 2025")

    # ---- pools ----
    with st.expander("🎯 Reservas de vaga (pools)"):
        config["pools"] = st.multiselect("Quais reservas de vaga este concurso tem?",
                                          POOLS_DISPONIVEIS,
                                          default=config.get("pools", POOLS_DISPONIVEIS))

    # ---- desempate ----
    with st.expander("⚖️ Critérios de desempate (em ordem)"):
        st.caption("Aplicados em cascata: só olha o próximo critério se o anterior empatar.")
        desemp = config.get("desempate", [])
        for i, crit in enumerate(list(desemp)):
            c1, c2, c3 = st.columns([3, 3, 1])
            with c1:
                crit["tipo"] = st.selectbox(f"Critério {i+1}", TIPOS_DESEMPATE,
                                             index=TIPOS_DESEMPATE.index(crit.get("tipo", "maior_nota")),
                                             key=f"desemp_tipo_{i}")
            with c2:
                if crit["tipo"] == "maior_nota":
                    crit["campo"] = st.text_input("Campo (nome do critério de correção usado no motor)",
                                                   value=crit.get("campo", ""), key=f"desemp_campo_{i}")
            with c3:
                st.write("")
                if st.button("🗑️", key=f"desemp_del_{i}"):
                    desemp.pop(i)
                    st.rerun()
        if st.button("+ Adicionar critério de desempate"):
            desemp.append({"tipo": "maior_nota", "campo": ""})
            st.rerun()
        config["desempate"] = desemp

    # ---- etapas ----
    with st.expander("📋 Etapas do concurso (em ordem)", expanded=True):
        etapas = config.get("etapas", [])
        for i, et in enumerate(list(etapas)):
            st.markdown(f"**Etapa {i+1}**")
            c1, c2, c3 = st.columns([3, 2, 1])
            with c1:
                et["nome"] = st.text_input("Nome da etapa", value=et.get("nome", ""), key=f"et_nome_{i}")
            with c2:
                et["tipo"] = st.selectbox("Tipo", TIPOS_ETAPA,
                                           index=TIPOS_ETAPA.index(et.get("tipo", "pontuacao")),
                                           key=f"et_tipo_{i}")
            with c3:
                st.write("")
                if st.button("🗑️", key=f"et_del_{i}"):
                    etapas.pop(i)
                    st.rerun()

            if et["tipo"] in ("pontuacao", "classificatoria"):
                et["campos"] = st.text_input("Campos de nota (separados por vírgula, nomes usados no motor)",
                                              value=",".join(et.get("campos", [])), key=f"et_campos_{i}")
                if et["tipo"] == "pontuacao":
                    cc1, cc2, cc3 = st.columns(3)
                    with cc1:
                        usar_corte = st.checkbox("Tem nota mínima?", value="corte" in et, key=f"et_temcorte_{i}")
                    with cc2:
                        minimo = st.number_input("Nota mínima (total)", value=float((et.get("corte") or {}).get("nota_minima_total", 0)),
                                                  key=f"et_min_{i}") if usar_corte else None
                    with cc3:
                        nao_zerar = st.checkbox("Não pode zerar nenhum item?",
                                                 value=(et.get("corte") or {}).get("nao_zerar", False),
                                                 key=f"et_zerar_{i}") if usar_corte else False
                    et["corte"] = {"nota_minima_total": minimo, "nao_zerar": nao_zerar} if usar_corte else None
                else:
                    et["corte"] = None
            elif et["tipo"] == "binaria":
                cc1, cc2, cc3 = st.columns(3)
                with cc1:
                    et["valor_aprovado"] = st.text_input("Valor considerado aprovado", value=et.get("valor_aprovado", "DEFERIDA"), key=f"et_valor_{i}")
                with cc2:
                    et["elimina_do_concurso"] = st.checkbox("Reprovar aqui elimina do concurso?",
                                                             value=et.get("elimina_do_concurso", False), key=f"et_elim_{i}")
                with cc3:
                    et["gate_pool"] = st.selectbox("Gateia qual pool? (opcional)", [""] + POOLS_DISPONIVEIS,
                                                    index=([""] + POOLS_DISPONIVEIS).index(et.get("gate_pool", "") or ""),
                                                    key=f"et_gate_{i}") or None
            st.markdown("---")
        if st.button("+ Adicionar etapa"):
            etapas.append({"nome": "", "tipo": "pontuacao", "campos": []})
            st.rerun()
        config["etapas"] = etapas

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("💾 Salvar configuração", type="secondary", use_container_width=True):
            if not config["nome"].strip():
                st.error("Dê um nome ao concurso antes de salvar.")
            else:
                for et in config["etapas"]:
                    if isinstance(et.get("campos"), str):
                        et["campos"] = [c.strip() for c in et["campos"].split(",") if c.strip()]
                r = salvar_config(gh_token, config["nome"].strip(), config)
                if r.status_code in (200, 201):
                    st.success("Configuração salva!")
                else:
                    st.error(f"Erro ao salvar: {r.status_code}")

    st.markdown("---")
    st.markdown("### 📤 Upload das planilhas")
    st.caption("Envie o arquivo de cada etapa configurada acima, na ordem. "
               "Para cada arquivo, escolha a aba e mapeie as colunas.")

    dados_etapas = st.session_state.setdefault("clf_uploads", {})

    # inscrições (fallback de pool)
    with st.expander("📂 Inscrições (usado como reserva de vaga quando o cargo não avança de etapa)"):
        arq_insc = st.file_uploader("Planilha de inscrições", type=["xlsx"], key="upl_inscricoes")
        if arq_insc:
            wb = ler_planilha(arq_insc.read())
            for pool in config.get("pools", []):
                sheet_options = [""] + wb.sheetnames
                sheet_sel = st.selectbox(f"Aba do pool '{pool}'", sheet_options, key=f"insc_sheet_{pool}")
                if sheet_sel:
                    cols = detectar_colunas(wb, sheet_sel, 2)
                    col_res = st.selectbox(f"Coluna de resultado ({pool})", [""] + [str(c) for c in cols], key=f"insc_col_{pool}")
                    dados_etapas.setdefault("inscricoes", {})[pool] = {"sheet": sheet_sel, "col_resultado": col_res, "wb": arq_insc.name}

    for i, et in enumerate(config.get("etapas", [])):
        nome_et = et.get("nome") or f"Etapa {i+1}"
        with st.expander(f"📂 {nome_et} ({et.get('tipo')})"):
            key_up = f"upl_etapa_{i}"
            arq = st.file_uploader(f"Planilha — {nome_et}", type=["xlsx"], key=key_up)
            if arq:
                wb = ler_planilha(arq.read())
                sheet_sel = st.selectbox("Aba", wb.sheetnames, key=f"et_sheet_{i}")
                cols = [str(c) for c in detectar_colunas(wb, sheet_sel, 2)]
                mapeamento = {}
                cc1, cc2, cc3, cc4 = st.columns(4)
                with cc1:
                    mapeamento["inscricao"] = st.selectbox("Coluna Inscrição", cols, key=f"et_map_insc_{i}")
                with cc2:
                    mapeamento["nome"] = st.selectbox("Coluna Candidato", cols, key=f"et_map_nome_{i}")
                with cc3:
                    mapeamento["nascimento"] = st.selectbox("Coluna Nascimento", [""] + cols, key=f"et_map_nasc_{i}")
                with cc4:
                    mapeamento["cargo"] = st.selectbox("Coluna Cargo", cols, key=f"et_map_cargo_{i}")

                if et.get("tipo") in ("pontuacao", "classificatoria"):
                    campos = et.get("campos", [])
                    if isinstance(campos, str):
                        campos = [c.strip() for c in campos.split(",") if c.strip()]
                    campos_map = {}
                    for campo in campos:
                        campos_map[campo] = st.selectbox(f"Coluna p/ '{campo}'", [""] + cols, key=f"et_map_campo_{i}_{campo}")
                    mapeamento["campos_map"] = campos_map
                    mapeamento["total"] = st.selectbox("Coluna Total da etapa", [""] + cols, key=f"et_map_total_{i}")
                    st.caption("Colunas de convocação por vaga nesta etapa (opcional — deixe em branco se não houver)")
                    pool_cols = {}
                    for pool in config.get("pools", []) + ["ampla"]:
                        pool_cols[pool] = st.selectbox(f"Coluna flag '{pool}'", [""] + cols, key=f"et_map_pool_{i}_{pool}")
                    mapeamento["pool_cols"] = pool_cols
                elif et.get("tipo") == "binaria":
                    mapeamento["resultado"] = st.selectbox("Coluna Resultado", [""] + cols, key=f"et_map_res_{i}")

                dados_etapas[f"etapa_{i}"] = {"wb": arq, "sheet": sheet_sel, "map": mapeamento}

    st.markdown("---")
    if st.button("⚙️ Gerar Classificação", type="primary", use_container_width=True):
        with st.spinner("Processando... isso pode levar um tempo em concursos grandes."):
            try:
                resultado_bytes, resumo = _processar(config, dados_etapas)
                st.success(f"✅ Classificação gerada! {resumo}")
                st.download_button(
                    "📥 Baixar Classificação Final (.xlsx)",
                    data=resultado_bytes,
                    file_name=f"CLASSIFICACAO_FINAL_{config['nome'].replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Erro ao processar: {e}")
                import traceback
                st.code(traceback.format_exc())
    st.stop()


def _rows_from_upload(wb, sheet, linha_cabecalho=2):
    ws = wb[sheet]
    header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i < linha_cabecalho:
            continue
        if i == linha_cabecalho:
            header = list(row)
            continue
        if row is None or row[0] is None:
            continue
        yield header, row


def _processar(config, dados_etapas):
    motor = MotorClassificacao({
        "pools": ["ampla"] + config.get("pools", []),
        "desempate": config.get("desempate", []),
    })

    # inscrições (fallback)
    insc_cfg = dados_etapas.get("inscricoes", {})
    if insc_cfg:
        pools_linhas = {}
        arq = None
        for pool, cfg in insc_cfg.items():
            pass  # arquivo já foi lido acima; refeito abaixo via session state se necessário

    n_candidatos_por_etapa = {}
    for i, et in enumerate(config.get("etapas", [])):
        chave = f"etapa_{i}"
        up = dados_etapas.get(chave)
        if not up:
            continue
        arq = up["wb"]
        arq.seek(0)
        wb = ler_planilha(arq.read())
        sheet = up["sheet"]
        mapa = up["map"]
        nome_et = et.get("nome") or chave

        linhas = []
        linhas_pool = []
        for header, row in _rows_from_upload(wb, sheet):
            def val(colname):
                if not colname or colname not in header:
                    return None
                return row[header.index(colname)]

            insc = val(mapa.get("inscricao"))
            if insc is None:
                continue
            nasc = val(mapa.get("nascimento"))
            if not isinstance(nasc, datetime.datetime):
                nasc = None
            item = {"inscricao": insc, "nome": val(mapa.get("nome")), "nascimento": nasc,
                    "cargo": val(mapa.get("cargo"))}

            if et.get("tipo") in ("pontuacao", "classificatoria"):
                valores = {campo: val(col) for campo, col in mapa.get("campos_map", {}).items()}
                item["valores"] = valores
                item["total"] = val(mapa.get("total"))
                linhas.append(item)
                flags = {}
                for pool, col in mapa.get("pool_cols", {}).items():
                    if col:
                        flags[pool] = val(col)
                if flags:
                    linhas_pool.append({**item, **flags})
            elif et.get("tipo") == "binaria":
                item["resultado"] = val(mapa.get("resultado"))
                linhas.append(item)

        if et.get("tipo") in ("pontuacao", "classificatoria"):
            campos = et.get("campos", [])
            if isinstance(campos, str):
                campos = [c.strip() for c in campos.split(",") if c.strip()]
            motor.carregar_etapa_pontuacao(nome_et, linhas, campos, corte=et.get("corte"),
                                            conta_no_total=True)
            if linhas_pool:
                mapa_cols = {pool: pool for pool in mapa.get("pool_cols", {}) if mapa["pool_cols"][pool]}
                motor.carregar_flags_pool(linhas_pool, mapa_cols)
        elif et.get("tipo") == "binaria":
            motor.carregar_etapa_binaria(nome_et, linhas,
                                          campo_valor_aprovado=et.get("valor_aprovado", "DEFERIDA"),
                                          elimina_do_concurso=et.get("elimina_do_concurso", False),
                                          gate_pool=et.get("gate_pool"))
        n_candidatos_por_etapa[nome_et] = len(linhas)

    ranking = motor.classificar(data_referencia_desempate=datetime.datetime.now())

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "CLASSIFICAÇÃO"
    header = ["INSCRIÇÃO", "CANDIDATO", "DATA DE NASCIMENTO", "CARGO"] + \
        [p.upper() for p in config.get("pools", [])] + \
        ["TOTAL GERAL"] + [f"CLASS. {p.upper()}" for p in ["ampla"] + config.get("pools", [])]
    ws_out.append(header)

    todos = sorted(motor.candidatos.values(), key=lambda c: (c.cargo, -c.total_geral))
    classes = {}  # (cargo,pool) -> {inscricao: rank}
    for cargo, pools in ranking.items():
        for pool, lista in pools.items():
            for idx, c in enumerate(lista, 1):
                classes[(cargo, pool, c.inscricao)] = idx

    total = 0
    for c in todos:
        if c.eliminado:
            continue
        linha = [c.inscricao, c.nome.strip(), c.nascimento, c.cargo]
        for p in config.get("pools", []):
            linha.append("SIM" if c.pools.get(p) else "-")
        linha.append(c.total_geral)
        for p in ["ampla"] + config.get("pools", []):
            linha.append(classes.get((c.cargo, p, c.inscricao), ""))
        ws_out.append(linha)
        total += 1

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    resumo = f"{total} candidatos classificados. Por etapa: {n_candidatos_por_etapa}"
    return buf.getvalue(), resumo
