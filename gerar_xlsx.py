"""
Gera cronograma IBGP no formato oficial (igual ao template Betim)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

DIAS_PT = ['Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira','Sexta-feira','Sábado','Domingo']

# Tarefas em negrito (publicação, períodos, provas, resultados pós-recurso, classificação, homologação)
BOLD_KEYWORDS = [
    'PUBLICAÇÃO','PERÍODO','PROVA OBJETIVA','PROVA DISCURSIVA','GABARITO PRELIMINAR',
    'RESULTADO PÓS-RECURSO','CLASSIFICAÇÃO FINAL','HOMOLOGAÇÃO','RESULTADO PRELIMINAR DO CURSO',
    'CLASSIFICAÇÃO PRELIMINAR','CDI','COMPROVANTE DEFINITIVO','REALIZAÇÃO','CURSO DE FORMAÇÃO'
]

def is_bold(atividade):
    a = atividade.upper()
    return any(k in a for k in BOLD_KEYWORDS)

def fmt_data(ini, fim):
    """Formata data no padrão IBGP: '02/01 a 06/01/2026' ou '27/10/2025'"""
    if ini.date() == fim.date():
        return ini.strftime('%d/%m/%Y')
    # Mesmo ano e mês
    if ini.year == fim.year:
        return f"{ini.strftime('%d/%m')} a {fim.strftime('%d/%m/%Y')}"
    return f"{ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"

def dia_semana(d):
    return DIAS_PT[d.weekday()]

def thin():
    return Side(border_style='thin', color='000000')

def gerar_xlsx_ibgp(nome_concurso, tipo, tarefas):
    """
    tarefas: lista de dicts com {seq, atividade, data_inicio (date), data_fim (date)}
    Retorna bytes do xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'CRONOGRAMA IBGP'

    # Dimensões das colunas (igual ao template)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22

    # ── Linha 1: título do concurso ──────────────────────────────
    ws.row_dimensions[1].height = 50
    ws.merge_cells('B1:C1')
    c1 = ws['B1']
    c1.value = nome_concurso
    c1.font = Font(name='Calibri', bold=True, size=12)
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Linha 2: "CRONOGRAMA PRELIMINAR" ────────────────────────
    ws.merge_cells('B2:C2')
    c2 = ws['B2']
    c2.value = 'CRONOGRAMA PRELIMINAR'
    c2.font = Font(name='Calibri', bold=True, size=12)
    c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Linha 3: cabeçalho da tabela ─────────────────────────────
    ws.row_dimensions[3].height = 20
    borda_hdr = Border(top=thin(), bottom=thin(), left=thin(), right=thin())
    for col, lbl in [('A','SEQ#'), ('B','Atividade'), ('C','Data'), ('D','Dia da Semana')]:
        cell = ws[f'{col}3']
        cell.value = lbl
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borda_hdr

    # ── Dados ─────────────────────────────────────────────────────
    borda_bd = Border(top=thin(), bottom=thin(), left=thin(), right=thin())
    borda_bd_sem_left = Border(top=thin(), bottom=thin(), right=thin())

    for i, t in enumerate(tarefas):
        row = i + 4
        ws.row_dimensions[row].height = 16.8
        ini = t['data_inicio']
        fim = t['data_fim']
        bold = is_bold(t['atividade'])

        # Col A: SEQ#
        ca = ws[f'A{row}']
        ca.value = t['seq']
        ca.font = Font(name='Times New Roman', size=12)
        ca.alignment = Alignment(horizontal='center', vertical='center')

        # Col B: Atividade
        cb = ws[f'B{row}']
        # Capitaliza igual ao template (primeira letra maiúscula, resto minúsculo por palavra-chave)
        cb.value = t['atividade'].capitalize() if not bold else t['atividade'].title().replace(' De ',' de ').replace(' Da ',' da ').replace(' Do ',' do ').replace(' E ',' e ').replace(' A ',' a ').replace(' O ',' o ').replace(' Para ',' para ').replace(' Dos ',' dos ').replace(' Das ',' das ').replace(' Com ',' com ').replace(' Contra ',' contra ').replace(' Se ',' se ')
        cb.font = Font(name='Calibri', bold=bold, size=11)
        cb.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cb.border = borda_bd

        # Col C: Data
        cc = ws[f'C{row}']
        cc.value = fmt_data(ini, fim)
        cc.font = Font(name='Calibri', bold=bold, size=11)
        cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cc.border = borda_bd

        # Col D: Dia da Semana
        cd = ws[f'D{row}']
        if ini.date() == fim.date():
            cd.value = dia_semana(ini)
        else:
            cd.value = ''
        cd.font = Font(name='Calibri', size=11)
        cd.alignment = Alignment(horizontal='center', vertical='center')
        cd.border = borda_bd_sem_left

    # ── Rodapé ────────────────────────────────────────────────────
    rodape_row = len(tarefas) + 4
    ws.merge_cells(f'B{rodape_row}:C{rodape_row}')
    cr = ws[f'B{rodape_row}']
    cr.value = 'Datas passíveis de alteração, acompanhe com frequência.\nTodos os resultados serão publicados após as 20h.'
    cr.font = Font(name='Calibri', italic=True, size=11)
    cr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[rodape_row].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

